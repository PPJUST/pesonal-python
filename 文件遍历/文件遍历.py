import os
from openpyxl import Workbook
import tkinter
from tkinter import filedialog
from openpyxl.styles import Font
from PySide2.QtCore import QFile
from PySide2.QtWidgets import QApplication,QMessageBox
from PySide2.QtUiTools import QUiLoader

# 导入ui
class Travelui:
    def __init__(self):
        super(Travelui, self).__init__()

        # 设置UI文件只读
        qfile = QFile("ui.ui")
        qfile.open(QFile.ReadOnly)
        qfile.close()

        # 加载UI
        self.ui = QUiLoader().load(qfile)

        # 初始化要用的变量
        self.file_names = []
        self.file_paths = []
        self.path_travel = "输入路径，或者点击按钮"
        self.path_result = os.getcwd()  # 取得当前路径

        # 绑定槽函数
        self.ui.button_path_travel.clicked.connect(self.path_get_button)   # 遍历文件选择路径按钮
        self.ui.button_result.clicked.connect(self.path_result_button)     # 结果路径按钮
        self.ui.button_start.clicked.connect(self.start_check)     # 开始遍历按钮
        self.ui.button_open_result.clicked.connect(self.open_result)    # 打开结果按钮
        self.ui.button_quit.clicked.connect(self.click_quit)

    # 设置ui的槽函数
    def ask_path(self):
        """弹窗选择文件夹路径"""
        root = tkinter.Tk()
        root.withdraw()  # 用来隐藏窗口
        path = filedialog.askdirectory()  # 选择路径弹窗
        return path

    def path_get_button(self):
        """遍历路径按钮"""
        self.path_travel = self.ask_path()
        self.ui.line_edit_path_travel.setText(self.path_travel) # 结果显示在文本框

    def path_result_button(self):
        """结果保存按钮"""
        self.path_result = self.ask_path()
        self.ui.line_edit_result.setText(self.path_result)  # 结果显示在文本框

    def start_check(self):
        """判断遍历路径是否为空"""
        if self.ui.line_edit_path_travel.text() == "输入路径，或者点击按钮":
            root = tkinter.Tk()
            root.withdraw()  # 用来隐藏窗口
            tkinter.messagebox.showwarning(title='注意', message='请选择需要遍历的文件夹路径')
        else:
            self.start_travel()

    def start_travel(self):
        """执行遍历操作"""
        # 初始化要用的变量
        self.file_names = []
        self.file_paths = []
        self.dir_path = []

        path_travel_result = os.walk(self.path_travel)  # 遍历后的文件路径
        for dirpath, dirnames, filenames in path_travel_result:  # for循环遍历
            for filename in filenames:
                self.file_names.append(filename)  # 添加文件名
                self.file_paths.append(os.path.join(dirpath, filename))  # 添加合并后的文件名和文件路径
                self.dir_path.append(dirpath)
        self.xlsx_mod() # 执行xlsx相关操作

    def xlsx_mod(self):
        """执行xlsx相关操作"""
        wb = Workbook()
        ws = wb.active
        ws.delete_cols(1), ws.delete_cols(1), ws.delete_cols(1)  # 删除3次第一列，清除原有数据
        ws.cell(1, 3).value, ws.cell(1, 1).value, ws.cell(1, 2).value = '文件完整路径', '文件名', '文件路径'  # 创建标题
        if self.ui.check_box_hyperlink.isChecked() == True:     # 判断超链接选项是否选中
            for linenumber in range(0, len(self.file_names) - 1):  # 按行写入数据
                ws.cell(linenumber + 2, 1).value = self.file_names[linenumber]  # 写入文件名
                ws.cell(linenumber + 2, 2).value = self.dir_path[linenumber]  # 写入文件路径
                ws.cell(linenumber + 2, 3).value = self.file_paths[linenumber]  # 写入文件完整路径
                ws.cell(linenumber + 2, 3).hyperlink = self.file_paths[linenumber]
                ws.cell(linenumber + 2, 3).font = Font(color="0000FF")
        else:
            for linenumber in range(0, len(self.file_names) - 1):  # 按行写入数据
                ws.cell(linenumber + 2, 1).value = self.file_names[linenumber]  # 写入文件名
                ws.cell(linenumber + 2, 2).value = self.dir_path[linenumber]  # 写入文件路径
                ws.cell(linenumber + 2, 3).value = self.file_paths[linenumber]  # 写入文件完整路径
        wb.save(f"{self.path_result}/文件遍历.xlsx")  # 保存表格
        tkinter.messagebox.showinfo(title='提示', message='已完成遍历操作，保存结果为"文件遍历.xlsx"')

    def open_result(self):
        """打开结果文件"""
        os.startfile(f"{self.path_result}/文件遍历.xlsx")

    def click_quit(self):
        quit()

def main():
    app = QApplication()
    travelui = Travelui()
    travelui.ui.show()
    app.exec_()

if __name__ == "__main__":
    main()