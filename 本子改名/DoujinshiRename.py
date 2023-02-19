import os
from openpyxl import Workbook
from openpyxl import load_workbook
import tkinter
from tkinter import filedialog
from PySide2.QtCore import QFile
from PySide2.QtWidgets import QApplication
from PySide2.QtUiTools import QUiLoader
import datetime
import re
import sys
import time
import random
import requests


# 导入ui
class DoujinshiRename:
    def __init__(self):
        super(DoujinshiRename, self).__init__()

        # 设置UI文件只读
        qfile = QFile("ui.ui")
        qfile.open(QFile.ReadOnly)
        qfile.close()

        # 加载UI
        self.ui = QUiLoader().load(qfile)

        # 初始化要用的变量
        self.read_config()  # 初始显示配置文件
        self.path = ""

        # 配置信号连接
        self.ui.button_path_get.clicked.connect(self.path_get_button)
        self.ui.button_path_open.clicked.connect(self.open_path_button)
        self.ui.button_open_help.clicked.connect(self.open_help_button)
        self.ui.button_check.clicked.connect(self.run_rename_check)
        self.ui.button_open_check.clicked.connect(self.open_check)
        self.ui.button_quit.clicked.connect(self.quit_botton)
        self.ui.button_start.clicked.connect(self.start_standard)
        self.ui.button_update_config.clicked.connect(self.update_config)
        self.ui.text_info.textChanged.connect(self.scroll)
        self.ui.button_back.clicked.connect(self.back_standard)
        self.ui.button_spiders_original.clicked.connect(self.spiders_original_check)

    def read_config(self):
        """读取配置文件，分配类"""
        # 首先对配置文件进行去空行
        with open("config.json", encoding="utf-8") as cr:
            crr = cr.readlines()
        with open("config.json", "w", encoding="utf-8") as cw:
            for line in crr:
                if line.split():
                    cw.write(line)

        config_all = []  # 设置个空列表
        with open("config.json", encoding="utf-8") as cj:
            config_all = cj.read().splitlines()
            line_row = 0    # 设置行，确定匹配到的关键词在第几行
            key_model, key_market, key_original, key_localization, key_other = 0, 0, 0, 0, 0
            for line in config_all:
                line_row += 1   # 每次循环+1行
                if line == "#model\n":
                    key_model = line_row
                if line == "#market":
                    key_market = line_row
                if line == "#original":
                    key_original = line_row
                if line == "#localization":
                    key_localization = line_row
                if line == "#other":
                    key_other = line_row
            self.config_model = config_all[key_model+1:key_market-1]
            self.config_market = list(tuple(config_all[key_market:key_original - 1]))   # 先转元组再转列表，完成去重
            self.config_original = list(tuple(config_all[key_original:key_localization - 1]))
            self.config_localization = list(tuple(config_all[key_localization:key_other - 1]))
            self.config_other = list(tuple(config_all[key_other:]))
            self.show_config()  # 执行方法

    def show_config(self):
        """显示配置文件中的关键词"""
        self.ui.line_edit_model.setText("".join(self.config_model))
        self.ui.text_edit_market.setText("\n".join(self.config_market))     # setText + \n.join组合列表完成多行显示列表中的元素
        self.ui.text_edit_original.setText("\n".join(self.config_original))
        self.ui.text_edit_localization.setText("\n".join(self.config_localization))
        self.ui.text_edit_other.setText("\n".join(self.config_other))
        self.ui.text_info.insertPlainText("\n" + self.get_time() + "程序初始化完成，已读取配置文件")

    def update_config(self):
        """更新配置文件"""
        import shutil
        newname = "config_backup " + str(datetime.datetime.now())[:-7].replace(':', '.') + ".json"
        p = os.getcwd() # 获取当前路径
        shutil.copy2("config.json", p + "\\config_backup\\" + newname)     # 调用shutil模块的方法复制文件
        self.ui.text_info.insertPlainText("\n" + self.get_time() + "已备份配置文件")

        # 设置临时变量，用于临时存放新的关键词（显示在文本框里的文本）
        new_config_model = self.ui.line_edit_model.text().split()
        new_config_market = self.ui.text_edit_market.toPlainText().split("\n")      # 先读取文本，然后split+换行符组成正常的列表
        new_config_original = self.ui.text_edit_original.toPlainText().split("\n")
        new_config_localization = self.ui.text_edit_localization.toPlainText().split("\n")
        new_config_other = self.ui.text_edit_other.toPlainText().split("\n")
        with open("config.json", "w", encoding="utf-8") as cw:      # 打开文件，重新写入5组关键词
            new_config = ["#model"] + new_config_model + \
                         ["#market"] + list(set(new_config_market)) + \
                         ["#original"] + list(set(new_config_original)) + \
                         ["#localization"] + list(set(new_config_localization)) + \
                         ["#other"] + list(set(new_config_other))     # 转集合再转列表完成去重，然后重组全部的关键词
            for line in new_config:
                cw.write(line + "\n")
        self.ui.text_info.insertPlainText("\n" + self.get_time() + "已更新配置文件")
        # 更新配置文件后重新读取
        self.read_config()

    def path_get_button(self):
        """选取路径文件夹按钮弹窗"""
        root = tkinter.Tk()
        root.withdraw()  # 用来隐藏窗口
        self.path = filedialog.askdirectory()
        self.ui.line_edit_path.setText(self.path)  # 结果显示在文本框

    def open_path_button(self):
        """打开文件夹按钮"""
        os.startfile(self.path)
        self.ui.text_info.insertPlainText("\n" + self.get_time() + "打开文件夹")

    def open_help_button(self):
        """打开帮助文件"""
        os.startfile("说明文档.png")
        self.ui.text_info.insertPlainText("\n" + self.get_time() + "打开说明文档")

    def get_time(self):
        """获取当前时间"""
        tm = str(datetime.datetime.now())[:-7].replace(':', '.') + ":"
        return tm

    """改名主逻辑代码"""
    """模块1
    新建xlsx >>> 遍历文件名 >>> 保存到xlsx"""
    def xlsx_create(self):
        """新建工作簿并且写入原文件名"""
        wb = Workbook()
        try:
            wb.save("改名测试.xlsx")  # 创建xlsx
        except PermissionError:
            tkinter.messagebox.showerror(title='出错了！', message='请关闭改名测试文件后重试')
        wb = load_workbook("改名测试.xlsx")  # 加载xlsx
        ws = wb.active
        ws["A1"] = "原文件名"
        paths = os.listdir(self.path)
        for i in range(2, len(paths) + 2):  # 将遍历得到的文件名写入xlsx
            ai = "A" + str(i)
            ws[ai] = paths[i - 2]
        wb.save("改名测试.xlsx")

    """模块2
    区分括号内外元素 >>> 
    括号内元素 >>> 匹配txt中的关键词进行分类
    括号外元素 >>> 判断是压缩包还是文件夹 >>> 去除文件两端多余空格"""
    def assort(self, input_filename):
        """文件名内元素分类"""
        # 先在函数内初始化要用的变量
        self.filename_standard = ""    # 标准化后的文件名
        self.filename_inbrac = []  # 用于存放括号里的内容
        self.filename_outbrac = ""  # 用于存放括号外的内容
        self.item_market = []  # 分类：即卖会名
        self.item_original = []  # 分类：原作名
        self.item_localization = []  # 分类：汉化组
        self.item_other = []  # 分类：其他信息
        self.standard_model = self.ui.line_edit_model.text().split("@")    # 切割关键词组成模型
        filename_usetomod = input_filename

        # 尝试区分压缩包
        zip_check = False
        if os.path.splitext(filename_usetomod)[1].upper() in ['.ZIP', '.RAR', '.7Z']:    # 分割文件名后转大写，然后找压缩包后缀
            zip_check = True

        # 区分括号内外元素
        self.filename_inbrac.extend(re.findall(r"\[.*?\]|\(.*?\)|（.*?）|【.*?】", filename_usetomod))  # 正则提取括号内容，extend拼接列表
        self.filename_outbrac = re.sub(r"\[.*?\]|\(.*?\)|（.*?）|【.*?】", "", filename_usetomod)  # 正则删除括号内容

        # 分类括号内容
        # 判断括号元素是否包含关键词：即卖会名
        for line in self.config_market:
            for ele in self.filename_inbrac:
                if ele.find(line) != -1:  # find()方法不为-1说明匹配成功
                    self.item_market.append(ele)  # 将匹配成功的元素转移到对应变量
                    self.filename_inbrac.remove(ele)  # 删除已转移的括号内容
        # 判断括号元素是否包含关键词：原作名
        for line in self.config_original:  # 判断括号元素是否包含关键词：即卖会名
            for ele in self.filename_inbrac:
                if ele.find(line) != -1:  # find()方法不为-1说明匹配成功
                    self.item_original.append(ele)  # 将匹配成功的元素转移到对应变量
                    self.filename_inbrac.remove(ele)  # 删除已转移的括号内容
        # 判断括号元素是否包含关键词：汉化组
        for line in self.config_localization:  # 判断括号元素是否包含关键词：即卖会名
            for ele in self.filename_inbrac:
                if ele.find(line) != -1:  # find()方法不为-1说明匹配成功
                    self.item_localization.append(ele)  # 将匹配成功的元素转移到对应变量
                    self.filename_inbrac.remove(ele)  # 删除已转移的括号内容
        # 判断括号元素是否包含关键词：其他信息
        for line in self.config_other:  # 判断括号元素是否包含关键词：即卖会名
            for ele in self.filename_inbrac:
                if ele.find(line) != -1:  # find()方法不为-1说明匹配成功
                    self.item_other.append(ele)  # 将匹配成功的元素转移到对应变量
                    self.filename_inbrac.remove(ele)  # 删除已转移的括号内容

        # 重组文件名，设置分类
        filename_usetorecomb = input_filename
        remove_brac = self.item_market + self.item_original + self.item_localization + self.item_other    # 整合需要删除的分类元素
        for i in remove_brac:
            filename_usetorecomb = filename_usetorecomb.replace(i, "")  # 替换相关括号元素
            filename_usetorecomb = filename_usetorecomb.strip()  # 清除两端多余空格
            filename_usetorecomb = filename_usetorecomb.replace("  ", " ")  # 2个空格变1个空额清除多余空格
        if zip_check:  # 上述方法无法清除带后缀的文件 后缀与文件名之间的空格
            if filename_usetorecomb.find(" .", -5) != -1:
                filename_usetorecomb = filename_usetorecomb[:-5] + filename_usetorecomb[-5:].replace(" .", ".")  # 切片替换 .去除多余空格

        # 对照标准化文件名格式，对文件名进行元素的删改
        if zip_check == True:
            filename_usetorecomb_delzip = os.path.splitext(filename_usetorecomb)[0]    # splitext分离后缀
            filename_usetorecomb_suffix = os.path.splitext(filename_usetorecomb)[1]    # splitext提取后缀
            for i in self.standard_model:
                if i == "[社团名(作者名)]标题":
                    self.filename_standard = self.filename_standard + " " + filename_usetorecomb_delzip
                elif i == "(原作名)":
                    self.filename_standard = self.filename_standard + " " + " ".join(self.item_original)
                elif i == "(即卖会名)":
                    self.filename_standard = self.filename_standard + " " + " ".join(self.item_market)
                elif i == "[汉化]":
                    self.filename_standard = self.filename_standard + " " + " ".join(self.item_localization)
                elif i == "[其他信息]":
                    self.filename_standard = self.filename_standard + " " + " ".join(self.item_other)
            self.filename_standard = self.filename_standard + filename_usetorecomb_suffix
        else:
            for i in self.standard_model:
                if i == "[社团名(作者名)]标题":
                    self.filename_standard = self.filename_standard + " " + filename_usetorecomb
                elif i == "(原作名)":
                    self.filename_standard = self.filename_standard + " " + " ".join(self.item_original)
                elif i == "(即卖会名)":
                    self.filename_standard = self.filename_standard + " " + " ".join(self.item_market)
                elif i == "[汉化]":
                    self.filename_standard = self.filename_standard + " " + " ".join(self.item_localization)
                elif i == "[其他信息]":
                    self.filename_standard = self.filename_standard + " " + " ".join(self.item_other)

        # 上述执行完成后可能有多余空格，继续用上面方法去除空格
        self.filename_standard = self.filename_standard.strip()  # 清除两端多余空格
        self.filename_standard = self.filename_standard.replace("  ", " ")  # 2个空格变1个空额清除多余空格
        if self.filename_standard.find(" .", -5) != -1:
            self.filename_standard = self.filename_standard[:-5] + self.filename_standard[-5:].replace(" .", ".")  # 切片替换 .去除多余空格

    """模块3
    将已经处理过的元素、文件名写入xlsx"""

    def run_rename_check(self):
        """重组文件名，设置分类"""
        # 数据写入xlsx
        self.xlsx_create()      # 首先执行之前编写的创建xlsx的函数
        wb = load_workbook("改名测试.xlsx")
        ws = wb.active
        max_row = ws.max_row
        for i in range(2, max_row + 2 - 1):
            ai = "A" + str(i)
            filename_oringinal = ws[ai].value  # 赋值于变量，用于后续操作
            self.assort(filename_oringinal)  # 执行模块2进行分类
            ws["B1"], ws["C1"], ws["D1"], ws["E1"], ws["F1"] = "标准化文件名", "即卖会名", "原作名", "汉化", "其他信息"    # 添加标题
            bi, ci, di, ei, fi = "B" + str(i), "C" + str(i), "D" + str(i), "E" + str(i), "F" + str(i)    # 建立单元格编号
            if filename_oringinal == self.filename_standard:  # 如果新文件名和原文件名相同，清空数据
                ws[bi], ws[ci], ws[di], ws[ei], ws[fi] = "", "", "", "", ""
            else:  # 如果不同，则写入数据
                ws[bi] = self.filename_standard
                ws[ci] = ",".join(self.item_market)
                ws[di] = ",".join(self.item_original)
                ws[ei] = ",".join(self.item_localization)
                ws[fi] = ",".join(self.item_other)

        # 执行查重操作，如果重复则添加前缀【】
        dup_check = []
        for i in range(2, max_row + 2 - 1):
            ai = "A" + str(i)
            dup_check.append(ws[ai].value)  # 先将所有原文件名添加到查重列表
        for i in range(2, max_row + 2 - 1):
            bi = "B" + str(i)
            dup_check.append(ws[bi].value)  # 逐行添加新文件名到查重列表
            if ws[bi].value == "":
                pass
            elif dup_check.count(ws[bi].value) > 1:  # 如果新文件名在查重列表里，添加前缀
                ws[bi] = "【重复" + str(dup_check.count(ws[bi].value) - 1) + "】" + ws[bi].value
            else:
                dup_check.append(ws[bi].value)
        wb.save("改名测试.xlsx")
        self.ui.text_info.insertPlainText("\n" + self.get_time() + "完成改名测试，已生成xlsx文件")

    def start_standard(self):
        """正式执行改名操作"""
        # 改名前先备份一次测试文件
        import shutil
        newname = "改名测试 " + str(datetime.datetime.now())[:-7].replace(':', '.') + ".xlsx"
        p = os.getcwd() # 获取当前路径
        shutil.copy2("改名测试.xlsx", p + "\\config_backup\\" + newname)     # 调用shutil模块的方法复制文件
        self.ui.text_info.insertPlainText("\n" + self.get_time() + "已备份测试文件")

        n = 0  # 用于统计改名操作的次数
        wb = load_workbook("改名测试.xlsx")
        ws = wb.active
        max_row = ws.max_row  # 获取最大行数
        for i in range(2, max_row + 2):
            ai = "A" + str(i)
            bi = "B" + str(i)
            if ws[bi].value:  # 判断新文件名是否为空，不为空则执行改名
                os.rename(self.path + "/" + ws[ai].value, self.path + "/" + ws[bi].value)
                self.ui.text_info.insertPlainText("\n"*2 + self.get_time() + "执行改名： " + ws[ai].value + " >>> " + ws[bi].value)
                n += 1  # 统计操作次数
        self.ui.text_info.insertPlainText("\n"*2 + self.get_time() + "已完成全部改名操作，共执行" + str(n) + "次")

    def back_standard(self):
        """撤销改名"""
        n = 0  # 用于统计改名操作的次数
        wb = load_workbook("改名测试.xlsx")
        ws = wb.active
        max_row = ws.max_row  # 获取最大行数
        for i in range(2, max_row + 2):
            ai = "A" + str(i)
            bi = "B" + str(i)
            if ws[bi].value:  # 判断新文件名是否为空，不为空则执行改名
                os.rename(self.path + "/" + ws[bi].value, self.path + "/" + ws[ai].value)
                self.ui.text_info.insertPlainText(
                    "\n" * 2 + self.get_time() + "撤回改名： " + ws[bi].value + " >>> " + ws[ai].value)
                n += 1  # 统计操作次数
        self.ui.text_info.insertPlainText("\n" * 2 + self.get_time() + "已撤回改名，共执行" + str(n) + "次")

    def open_check(self):
        """打开测试文件"""
        os.startfile("改名测试.xlsx")
        self.ui.text_info.insertPlainText("\n" + self.get_time() + "打开测试文件")

    def quit_botton(self):
        sys.exit(1)

    def scroll(self):
        """文本框下拉到底"""
        self.ui.text_info.verticalScrollBar().setValue(self.ui.text_info.verticalScrollBar().maximum())

    def spiders_original_check(self):
        """爬虫添加密码"""
        password = tkinter.simpledialog.askstring(title='输入密码', prompt='输入密码：')
        if password == "给爷爬":
            self.spiders_original(1, 900)
        else:
            tkinter.messagebox.showwarning(title='密码错误', message='请勿使用')

    def spiders_original(self, min_page, max_page):
        """爬取原作名"""
        user_agents = [
            "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; AcooBrowser; .NET CLR 1.1.4322; .NET CLR 2.0.50727)",
            "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0; Acoo Browser; SLCC1; .NET CLR 2.0.50727; Media Center PC 5.0; .NET CLR 3.0.04506)",
            "Mozilla/4.0 (compatible; MSIE 7.0; AOL 9.5; AOLBuild 4337.35; Windows NT 5.1; .NET CLR 1.1.4322; .NET CLR 2.0.50727)",
            "Mozilla/5.0 (Windows; U; MSIE 9.0; Windows NT 9.0; en-US)",
            "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Win64; x64; Trident/5.0; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 2.0.50727; Media Center PC 6.0)",
            "Mozilla/5.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0; WOW64; Trident/4.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 1.0.3705; .NET CLR 1.1.4322)",
            "Mozilla/4.0 (compatible; MSIE 7.0b; Windows NT 5.2; .NET CLR 1.1.4322; .NET CLR 2.0.50727; InfoPath.2; .NET CLR 3.0.04506.30)",
            "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN) AppleWebKit/523.15 (KHTML, like Gecko, Safari/419.3) Arora/0.3 (Change: 287 c9dfb30)",
            "Mozilla/5.0 (X11; U; Linux; en-US) AppleWebKit/527+ (KHTML, like Gecko, Safari/419.3) Arora/0.6",
            "Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.8.1.2pre) Gecko/20070215 K-Ninja/2.1.1",
            "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9) Gecko/20080705 Firefox/3.0 Kapiko/3.0",
            "Mozilla/5.0 (X11; Linux i686; U;) Gecko/20070322 Kazehakase/0.4.5",
            "Mozilla/5.0 (X11; U; Linux i686; en-US; rv:1.9.0.8) Gecko Fedora/1.9.0.8-1.fc10 Kazehakase/0.5.6",
            "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 Safari/535.11",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_3) AppleWebKit/535.20 (KHTML, like Gecko) Chrome/19.0.1036.7 Safari/535.20",
            "Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; fr) Presto/2.9.168 Version/11.52",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.122 Safari/537.36"
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.163 Safari/537.36"
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.106 Safari/537.36"
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:76.0) Gecko/20100101 Firefox/76.0"
            "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.132 Safari/537.36"
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.149 Safari/537.36"
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.122 Safari/537.36"
            "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:76.0) Gecko/20100101 Firefox/76.0"
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_5) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.1.2 Safari/605.1.15"
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_5) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.1.1 Safari/605.1.15"
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.106 Safari/537.36Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:78.0) Gecko/20100101 Firefox/78.0"
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.87 Safari/537.36"
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.106 Safari/537.36"
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_4) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.1 Safari/605.1.15"
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_4) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.1.2 Safari/605.1.15"
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
            "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:75.0) Gecko/20100101 Firefox/75.0"
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_4) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.1.1 Safari/605.1.15"
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.132 Safari/537.36"
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.149 Safari/537.36"
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.102 Safari/537.36"
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_5) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.1.3 Safari/605.1.15"
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:78.0) Gecko/20100101 Firefox/78.0"
            "Mozilla/5.0 (Windows NT 10.0; WOW64; Trident/7.0; rv:11.0) like Gecko"
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.163 Safari/537.36"
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:75.0) Gecko/20100101 Firefox/75.0"
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_4) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.1.4 Safari/605.1.15"
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.87 Safari/537.36"
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_4) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.1.3 Safari/605.1.15"
            "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.87 Safari/537.36"
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:75.0) Gecko/20100101 Firefox/75.0"
            "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0"
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_5) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.1 Safari/605.1.15"
        ]
        for page in range(min_page, max_page + 1):  # 设置for循环
            htmls = []  # 设置空列表，存放爬取的全部网页内容
            self.ui.text_info.insertPlainText("\n" * 2 + self.get_time() + "正在爬取第" + str(page) + "页")
            url = f'https://bangumi.tv/anime/browser?page={page}'  # 设置url
            response = requests.get(url, headers={'User-Agent': random.choice(user_agents)})
            response.encoding = "utf-8"
            if response.status_code == 200:  # 检查状态码
                content = response.text
            htmls = content.split("\n")  # 将爬虫结果用换行符分割为列表
            self.ui.text_info.insertPlainText("\n" * 2 + self.get_time() + "完成第" + str(page) + "页的爬取")
            # 进行正则操作
            results = []  # 设置空列表来存放最终原作名
            for i in htmls:
                if i.find('class="l"') != -1 and i.find("small") != -1:
                    try:
                        results.append(re.search(r'class="l">(.*?)</a>', i).group(1).strip())  # 正则提取关键词
                    except AttributeError:  # 如果正则没有提取则报错，则跳过
                        pass
                    try:
                        results.append(re.search(r'class="grey">(.*?)</small>', i).group(1).strip())
                    except AttributeError:
                        pass
            # 替换Windows文件名不能出现的字符
            for i in results:
                results.remove(i)
                newi = i.replace("/", " ")  # 替换Windows文件名不能出现的/
                results.append(newi)
            # 检查原作名长度，如果过短则添加括号防止匹配错误
            results_check = []  # 检查文件长度的空列表
            for i in results:
                if len(str(i)) <= 8:
                    results.remove(i)
                    results_check.append("(" + i + ")")
            results += results_check
            with open("原作爬取结果.txt", "a", encoding="utf-8") as rw:
                for i in results:
                    rw.write(i + "\n")
            self.ui.text_info.insertPlainText("\n" + self.get_time() + "已将结果写入文件，等待5秒后开始下一次爬取")
            time.sleep(5)
        self.ui.text_info.insertPlainText("\n" + self.get_time() + "完成全部爬取")


def main():
    app = QApplication()
    #app.setStyle('Fusion')
    doujinshirename = DoujinshiRename()
    doujinshirename.ui.show()
    app.exec_()

if __name__ == "__main__":
    main()
